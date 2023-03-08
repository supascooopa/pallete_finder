import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO


def color_excel_creator(color_list):
    wb = openpyxl.Workbook()
    ws = wb.active

    clean_color_list = [color.replace("#", "") for color in color_list]
    hex_rows = [[hex_code] for hex_code in color_list]

    for row in hex_rows:
        # appending hex color codes to Excel sheet
        ws.append(row)

    for row_num, hex_code in enumerate(clean_color_list, 1):
        # reading hex color codes from the list to apply that color on the cell with corresponding hex code
        color_fill = PatternFill(start_color=hex_code, end_color=hex_code, fill_type='solid')
        ws[f'A{row_num}'].fill = color_fill

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    # must return to the beginning
    virtual_workbook.seek(0)
    return virtual_workbook

